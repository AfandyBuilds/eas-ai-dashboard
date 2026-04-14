#!/usr/bin/env python3
"""
EAS AI Adoption — Tracker Sheet → SQL Sync Generator

Reads the EAS AI Adoption Weekly Tracker Excel and produces SQL INSERT/UPDATE
statements suitable for execution via Supabase MCP execute_sql.

Usage:
    python scripts/sync_tracker.py <path-to-tracker.xlsx> [--out <output-dir>]

Output:
    Generated SQL files in <output-dir>/  (default: scripts/sync_output/)
      - sync_copilot_users.sql
      - sync_tasks.sql
      - sync_projects.sql
      - sync_accomplishments.sql

Notes:
    - Uses ON CONFLICT for upsert logic (append/update, never delete)
    - Copilot users matched by LOWER(TRIM(email))
    - Tasks matched by sync_hash (MD5 of practice+employee+week+task)
    - Projects matched by LOWER(project_name) + project_code per practice
    - Accomplishments matched by sync_hash (MD5 of practice+title+date)
"""
import sys
import os
import hashlib
import argparse
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)


def escape_sql(val):
    """Escape a value for SQL string literal."""
    if val is None:
        return "NULL"
    s = str(val).strip()
    if not s or s.lower() == 'nan' or s.lower() == 'none':
        return "NULL"
    s = s.replace("'", "''")
    return f"'{s}'"


def to_num(val, default=0):
    """Convert to numeric, return default if not possible."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def to_int(val, default=0):
    """Convert to int."""
    return int(to_num(val, default))


def md5_hash(*parts):
    """Generate MD5 hash from parts for dedup."""
    combined = '|'.join(str(p).strip().lower() for p in parts if p)
    return hashlib.md5(combined.encode()).hexdigest()


def parse_date(val):
    """Try to parse a date value to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s.lower() in ('nan', 'none', 'nat'):
        return None
    # Try common formats
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def determine_quarter(date_str):
    """Given a date string YYYY-MM-DD, return quarter ID like Q1-2026."""
    if not date_str:
        return 'Q1-2026'  # default
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        q = (dt.month - 1) // 3 + 1
        return f"Q{q}-{dt.year}"
    except ValueError:
        return 'Q1-2026'


# ──────────────────────────────────────────────
# Practice sheet names → practice DB names
# ──────────────────────────────────────────────
PRACTICE_SHEET_MAP = {
    'BFSI': 'BFSI',
    'CES': 'CES',
    'ERP': 'ERP Solutions',
    'EPS': 'EPS',
    'EPCS': 'EPCS',
    'GRC': 'GRC',
}


def extract_copilot_users(wb):
    """Extract copilot users from 'Copliot User Access' sheet."""
    sheet_name = 'Copliot User Access'
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found, skipping copilot users.")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # Header: Practice, Resource Name, Resource Email ID, Skill, Remarks, Business line
    users = []
    for row in rows[1:]:  # skip header
        if not row or len(row) < 5:
            continue
        practice = str(row[0]).strip() if row[0] else None
        name = str(row[1]).strip() if row[1] else None
        email = str(row[2]).strip().lower() if row[2] else None
        skill = str(row[3]).strip() if row[3] else None
        remarks = str(row[4]).strip() if row[4] else 'access granted'
        business_line = str(row[5]).strip() if len(row) > 5 and row[5] else None

        if not email or not name or not practice:
            continue
        if email in ('nan', 'none', ''):
            continue

        # Normalize practice name
        practice_db = practice
        if practice in PRACTICE_SHEET_MAP:
            practice_db = PRACTICE_SHEET_MAP[practice]

        users.append({
            'practice': practice_db,
            'name': name,
            'email': email,
            'role_skill': skill,
            'status': remarks if remarks and remarks.lower() != 'nan' else 'access granted',
        })

    return users


def extract_tasks(wb):
    """Extract tasks from practice-specific sheets."""
    all_tasks = []

    for sheet_name, practice_db in PRACTICE_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping tasks.")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find the header row (contains 'Week #' or 'Week Start Date')
        header_idx = None
        for i, row in enumerate(rows):
            if row and any(str(c).strip().lower().startswith('week') for c in row if c):
                # Check if this looks like the data header
                row_str = [str(c).strip().lower() if c else '' for c in row]
                if 'employee name' in row_str or 'week #' in row_str:
                    header_idx = i
                    break

        if header_idx is None:
            print(f"  WARNING: Could not find header row in '{sheet_name}', skipping.")
            continue

        # Expected columns:
        # 0: Week #, 1: Week Start Date, 2: Week End Date, 3: Project Name,
        # 4: Project Code, 5: Employee Name, 6: Task / Activity, 7: Task Category,
        # 8: AI Tool Used, 9: Prompt / Approach, 10: Time Without AI (hrs),
        # 11: Time With AI (hrs), 12: Time Saved (hrs), 13: Efficiency Gain %,
        # 14: Quality Rating (1-5), 15: Status, 16: Notes / Learnings

        for row in rows[header_idx + 1:]:
            if not row or len(row) < 12:
                continue

            employee = str(row[5]).strip() if row[5] else None
            task_desc = str(row[6]).strip() if row[6] else None

            if not employee or not task_desc:
                continue
            if employee.lower() in ('nan', 'none', ''):
                continue

            week_num = to_int(row[0]) if row[0] else None
            week_start = parse_date(row[1])
            week_end = parse_date(row[2])
            project = str(row[3]).strip() if row[3] else None
            project_code = str(row[4]).strip() if row[4] else None
            category = str(row[7]).strip() if row[7] else ''
            ai_tool = str(row[8]).strip() if row[8] else ''
            prompt = str(row[9]).strip() if len(row) > 9 and row[9] else None
            time_without = to_num(row[10]) if len(row) > 10 else 0
            time_with = to_num(row[11]) if len(row) > 11 else 0
            quality = to_num(row[14]) if len(row) > 14 and row[14] else 0
            status = str(row[15]).strip() if len(row) > 15 and row[15] else 'Completed'
            notes = str(row[16]).strip() if len(row) > 16 and row[16] else None

            # Determine quarter from week_start date
            quarter_id = determine_quarter(week_start)

            # Generate sync hash for dedup
            sync = md5_hash(practice_db, employee, week_num or '', task_desc)

            all_tasks.append({
                'quarter_id': quarter_id,
                'practice': practice_db,
                'week_number': week_num,
                'week_start': week_start,
                'week_end': week_end,
                'project': project,
                'project_code': project_code,
                'employee_name': employee,
                'task_description': task_desc,
                'category': category,
                'ai_tool': ai_tool,
                'prompt_used': prompt,
                'time_without_ai': time_without,
                'time_with_ai': time_with,
                'quality_rating': quality,
                'status': status if status and status.lower() not in ('nan', 'none') else 'Completed',
                'notes': notes,
                'sync_hash': sync,
            })

    return all_tasks


def extract_projects(wb):
    """Extract projects from the Projects sheet."""
    if 'Projects' not in wb.sheetnames:
        print("  WARNING: 'Projects' sheet not found, skipping.")
        return []

    ws = wb['Projects']
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    # Layout: columns are grouped by practice in pairs (Project Name, Project Code, blank)
    # Row 0: practice group headers: 'ERP Projects', None, None, 'CES Projects', ...
    # Row 1: 'Project Name', 'Project Code', None, 'Project Name', 'Project Code', ...
    # Row 2+: data

    practice_groups = []
    header_row = rows[0]
    for i, cell in enumerate(header_row):
        if cell and 'projects' in str(cell).lower():
            practice_name = str(cell).replace('Projects', '').strip()
            practice_groups.append((i, practice_name))

    projects = []
    for start_col, practice_label in practice_groups:
        # Map practice label to DB name
        practice_db = None
        for key, val in PRACTICE_SHEET_MAP.items():
            if key.lower() in practice_label.lower() or practice_label.lower() in val.lower():
                practice_db = val
                break
        if not practice_db:
            # Try direct match
            if practice_label.upper() in PRACTICE_SHEET_MAP:
                practice_db = PRACTICE_SHEET_MAP[practice_label.upper()]
            else:
                practice_db = practice_label

        for row in rows[2:]:  # skip header rows
            if len(row) <= start_col:
                continue
            proj_name = row[start_col]
            proj_code = row[start_col + 1] if len(row) > start_col + 1 else None

            if not proj_name or str(proj_name).strip().lower() in ('nan', 'none', ''):
                continue

            proj_name_str = str(proj_name).strip()
            proj_code_str = str(proj_code).strip() if proj_code else None
            if proj_code_str and proj_code_str.lower() in ('nan', 'none'):
                proj_code_str = None

            sync = md5_hash(practice_db, proj_name_str, proj_code_str or '')

            projects.append({
                'practice': practice_db,
                'project_name': proj_name_str,
                'project_code': proj_code_str,
                'sync_hash': sync,
            })

    return projects


def extract_accomplishments(wb):
    """Extract accomplishments from 'AI Accomplishments' sheet."""
    sheet_name = 'AI Accomplishments'
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found, skipping.")
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (row with '#', 'Date', 'Practice', ...)
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip() == '#' for c in row if c):
            header_idx = i
            break

    if header_idx is None:
        print("  WARNING: Could not find header in AI Accomplishments.")
        return []

    # Columns (from inspection):
    # 0: #, 1: Date, 2: Practice, 3: Project, 4: Project Code,
    # 5: SPOC/Lead, 6: Employee(s) Involved, 7: Accomplishment Title,
    # 8: Accomplishment Details, 9: AI Tool Used, 10: Category,
    # 11: Before AI (Baseline), 12: After AI (Result), 13: Quantified Impact,
    # 14: Business Gains, 15: Cost to Client, 16: Effort Saved (hrs),
    # 17: Status, 18: Evidence / Links, 19: Management Notes

    accs = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 10:
            continue

        title = str(row[7]).strip() if row[7] else None
        practice = str(row[2]).strip() if row[2] else None

        if not title or not practice:
            continue
        if title.lower() in ('nan', 'none'):
            continue

        date_val = parse_date(row[1])
        quarter_id = determine_quarter(date_val)
        project = str(row[3]).strip() if row[3] else None
        project_code = str(row[4]).strip() if row[4] else None
        spoc = str(row[5]).strip() if row[5] else None
        employees = str(row[6]).strip() if row[6] else None
        details = str(row[8]).strip() if row[8] else None
        ai_tool = str(row[9]).strip() if row[9] else None
        category = str(row[10]).strip() if len(row) > 10 and row[10] else None
        before_baseline = str(row[11]).strip() if len(row) > 11 and row[11] else None
        after_result = str(row[12]).strip() if len(row) > 12 and row[12] else None
        quantified_impact = str(row[13]).strip() if len(row) > 13 and row[13] else None
        business_gains = str(row[14]).strip() if len(row) > 14 and row[14] else None
        cost = str(row[15]).strip() if len(row) > 15 and row[15] else 'Free of Cost'
        effort_saved = to_num(row[16]) if len(row) > 16 else None
        status = str(row[17]).strip() if len(row) > 17 and row[17] else 'Completed'
        evidence = str(row[18]).strip() if len(row) > 18 and row[18] else None
        notes = str(row[19]).strip() if len(row) > 19 and row[19] else None

        sync = md5_hash(practice, title, date_val or '')

        accs.append({
            'quarter_id': quarter_id,
            'practice': practice,
            'date': date_val,
            'project': project,
            'project_code': project_code,
            'spoc': spoc,
            'employees': employees,
            'title': title,
            'details': details,
            'ai_tool': ai_tool,
            'category': category,
            'before_baseline': before_baseline,
            'after_result': after_result,
            'quantified_impact': quantified_impact,
            'business_gains': business_gains,
            'cost': cost,
            'effort_saved': effort_saved,
            'status': status if status and status.lower() not in ('nan', 'none') else 'Completed',
            'evidence': evidence,
            'notes': notes,
            'sync_hash': sync,
        })

    return accs


# ──────────────────────────────────────────────
# SQL Generation
# ──────────────────────────────────────────────

def gen_copilot_users_sql(users):
    """Generate upsert SQL for copilot users."""
    lines = [
        "-- Copilot Users Sync (generated " + datetime.now().isoformat() + ")",
        "-- Upserts by LOWER(TRIM(email))",
        ""
    ]

    for u in users:
        email_lower = u['email'].lower().strip()
        lines.append(f"""INSERT INTO copilot_users (practice, name, email, role_skill, status, sync_source, last_synced_at)
VALUES ({escape_sql(u['practice'])}, {escape_sql(u['name'])}, {escape_sql(email_lower)}, {escape_sql(u['role_skill'])}, {escape_sql(u['status'])}, 'tracker_sync', now())
ON CONFLICT (email) DO UPDATE SET
  practice = EXCLUDED.practice,
  name = EXCLUDED.name,
  role_skill = COALESCE(EXCLUDED.role_skill, copilot_users.role_skill),
  status = EXCLUDED.status,
  sync_source = 'tracker_sync',
  last_synced_at = now();
""")

    return '\n'.join(lines)


def gen_tasks_sql(tasks):
    """Generate upsert SQL for tasks."""
    lines = [
        "-- Tasks Sync (generated " + datetime.now().isoformat() + ")",
        "-- Uses sync_hash for idempotent upsert",
        ""
    ]

    for t in tasks:
        lines.append(f"""INSERT INTO tasks (quarter_id, practice, week_number, week_start, week_end,
  project, project_code, employee_name, task_description, category, ai_tool,
  prompt_used, time_without_ai, time_with_ai, quality_rating, status, notes,
  source, sync_hash, approval_status)
VALUES ({escape_sql(t['quarter_id'])}, {escape_sql(t['practice'])}, {t['week_number'] if t['week_number'] else 'NULL'},
  {escape_sql(t['week_start'])}, {escape_sql(t['week_end'])},
  {escape_sql(t['project'])}, {escape_sql(t['project_code'])},
  {escape_sql(t['employee_name'])}, {escape_sql(t['task_description'])},
  {escape_sql(t['category'])}, {escape_sql(t['ai_tool'])},
  {escape_sql(t['prompt_used'])}, {t['time_without_ai']}, {t['time_with_ai']},
  {t['quality_rating']}, {escape_sql(t['status'])}, {escape_sql(t['notes'])},
  'tracker_sync', {escape_sql(t['sync_hash'])}, 'approved')
ON CONFLICT (sync_hash) WHERE sync_hash IS NOT NULL DO UPDATE SET
  time_without_ai = EXCLUDED.time_without_ai,
  time_with_ai = EXCLUDED.time_with_ai,
  quality_rating = EXCLUDED.quality_rating,
  status = EXCLUDED.status,
  notes = EXCLUDED.notes;
""")

    return '\n'.join(lines)


def gen_projects_sql(projects):
    """Generate upsert SQL for projects."""
    lines = [
        "-- Projects Sync (generated " + datetime.now().isoformat() + ")",
        "-- Uses sync_hash for idempotent upsert",
        ""
    ]

    for p in projects:
        lines.append(f"""INSERT INTO projects (practice, project_name, project_code, is_active, sync_hash)
VALUES ({escape_sql(p['practice'])}, {escape_sql(p['project_name'])}, {escape_sql(p['project_code'])}, true, {escape_sql(p['sync_hash'])})
ON CONFLICT (sync_hash) WHERE sync_hash IS NOT NULL DO UPDATE SET
  project_name = EXCLUDED.project_name,
  project_code = EXCLUDED.project_code;
""")

    return '\n'.join(lines)


def gen_accomplishments_sql(accs):
    """Generate upsert SQL for accomplishments."""
    lines = [
        "-- Accomplishments Sync (generated " + datetime.now().isoformat() + ")",
        "-- Uses sync_hash for idempotent upsert",
        ""
    ]

    for a in accs:
        lines.append(f"""INSERT INTO accomplishments (quarter_id, practice, date, project, project_code,
  spoc, employees, title, details, ai_tool, category, before_baseline, after_result,
  quantified_impact, business_gains, cost, effort_saved, status, evidence, notes,
  source, sync_hash, approval_status)
VALUES ({escape_sql(a['quarter_id'])}, {escape_sql(a['practice'])}, {escape_sql(a['date'])},
  {escape_sql(a['project'])}, {escape_sql(a['project_code'])},
  {escape_sql(a['spoc'])}, {escape_sql(a['employees'])},
  {escape_sql(a['title'])}, {escape_sql(a['details'])},
  {escape_sql(a['ai_tool'])}, {escape_sql(a['category'])},
  {escape_sql(a['before_baseline'])}, {escape_sql(a['after_result'])},
  {escape_sql(a['quantified_impact'])}, {escape_sql(a['business_gains'])},
  {escape_sql(a['cost'])}, {a['effort_saved'] if a['effort_saved'] else 'NULL'},
  {escape_sql(a['status'])}, {escape_sql(a['evidence'])}, {escape_sql(a['notes'])},
  'tracker_sync', {escape_sql(a['sync_hash'])}, 'approved')
ON CONFLICT (sync_hash) WHERE sync_hash IS NOT NULL DO UPDATE SET
  details = EXCLUDED.details,
  effort_saved = EXCLUDED.effort_saved,
  status = EXCLUDED.status,
  notes = EXCLUDED.notes;
""")

    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='Sync EAS Tracker Excel → SQL')
    parser.add_argument('xlsx', help='Path to the tracker Excel file')
    parser.add_argument('--out', default='scripts/sync_output', help='Output directory for SQL files')
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: File not found: {args.xlsx}")
        sys.exit(1)

    os.makedirs(args.out, exist_ok=True)

    print(f"Reading tracker: {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # Extract data
    print("\nExtracting copilot users...")
    users = extract_copilot_users(wb)
    print(f"  Found {len(users)} users")

    print("Extracting tasks...")
    tasks = extract_tasks(wb)
    print(f"  Found {len(tasks)} tasks")

    print("Extracting projects...")
    projects = extract_projects(wb)
    print(f"  Found {len(projects)} projects")

    print("Extracting accomplishments...")
    accs = extract_accomplishments(wb)
    print(f"  Found {len(accs)} accomplishments")

    wb.close()

    # Generate SQL files
    print(f"\nGenerating SQL files in {args.out}/")

    with open(os.path.join(args.out, 'sync_copilot_users.sql'), 'w', encoding='utf-8') as f:
        f.write(gen_copilot_users_sql(users))
    print(f"  [OK] sync_copilot_users.sql ({len(users)} records)")

    with open(os.path.join(args.out, 'sync_tasks.sql'), 'w', encoding='utf-8') as f:
        f.write(gen_tasks_sql(tasks))
    print(f"  [OK] sync_tasks.sql ({len(tasks)} records)")

    with open(os.path.join(args.out, 'sync_projects.sql'), 'w', encoding='utf-8') as f:
        f.write(gen_projects_sql(projects))
    print(f"  [OK] sync_projects.sql ({len(projects)} records)")

    with open(os.path.join(args.out, 'sync_accomplishments.sql'), 'w', encoding='utf-8') as f:
        f.write(gen_accomplishments_sql(accs))
    print(f"  [OK] sync_accomplishments.sql ({len(accs)} records)")

    print("\nDONE! Execute the SQL files via Supabase MCP execute_sql.")
    print("   Order: copilot_users → projects → tasks → accomplishments")


if __name__ == '__main__':
    main()
