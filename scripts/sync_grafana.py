#!/usr/bin/env python3
"""
EAS AI Adoption — Grafana IDE Usage → SQL Sync Generator

Reads Grafana Copilot IDE usage Excel exports and produces SQL UPDATE
statements to populate the ide_* columns on copilot_users.

The script:
  1. Reads one or more Grafana Excel files (daily per-user IDE metrics)
  2. Filters to EAS employees by matching user_login against copilot_users emails
  3. Aggregates daily data per user
  4. Generates SQL UPDATEs keyed on the username column in copilot_users

Usage:
    python scripts/sync_grafana.py <grafana1.xlsx> [grafana2.xlsx ...]
        [--emails <copilot_emails.txt>] [--out <output-dir>]

    If --emails is not provided, generates UPDATEs for ALL user logins found.
    To filter to EAS users only, provide a file with one email per line.

Output:
    scripts/sync_output/sync_grafana_ide.sql
"""
import sys
import os
import argparse
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl")
    sys.exit(1)


def escape_sql(val):
    if val is None:
        return "NULL"
    s = str(val).strip()
    if not s or s.lower() in ('nan', 'none'):
        return "NULL"
    s = s.replace("'", "''")
    return f"'{s}'"


def read_grafana_file(filepath):
    """
    Read a Grafana IDE usage Excel file and return list of daily records.
    
    Expected columns (from header row):
        day, user_login, user_initiated_interaction_count,
        code_generation_activity_count, code_acceptance_activity_count,
        used_agent, used_chat, loc_suggested_to_add_sum,
        loc_suggested_to_delete_sum, loc_added_sum, loc_deleted_sum,
        [used_cli, used_copilot_coding_agent, LOC Add Actual]
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    # Parse header
    header = [str(h).strip().lower() if h else '' for h in rows[0]]

    # Map expected column names to indices
    col_map = {}
    expected = {
        'day': 'day',
        'user_login': 'user_login',
        'user_initiated_interaction_count': 'interactions',
        'code_generation_activity_count': 'code_gen',
        'code_acceptance_activity_count': 'code_accept',
        'used_agent': 'used_agent',
        'used_chat': 'used_chat',
        'loc_suggested_to_add_sum': 'loc_suggested',
        'loc_suggested_to_delete_sum': 'loc_suggested_del',
        'loc_added_sum': 'loc_added',
        'loc_deleted_sum': 'loc_deleted',
    }

    for i, h in enumerate(header):
        for excel_name, our_name in expected.items():
            if h == excel_name:
                col_map[our_name] = i

    records = []
    for row in rows[1:]:
        if not row:
            continue

        user_login = str(row[col_map.get('user_login', 1)]).strip().lower() if row[col_map.get('user_login', 1)] else None
        if not user_login or user_login in ('nan', 'none', ''):
            continue

        day = row[col_map.get('day', 0)]
        if isinstance(day, datetime):
            day_str = day.strftime('%Y-%m-%d')
        elif day:
            day_str = str(day).strip()
        else:
            day_str = None

        def safe_int(idx_key, default=0):
            idx = col_map.get(idx_key)
            if idx is None or idx >= len(row) or row[idx] is None:
                return default
            try:
                return int(float(str(row[idx])))
            except (ValueError, TypeError):
                return default

        def safe_bool(idx_key):
            idx = col_map.get(idx_key)
            if idx is None or idx >= len(row) or row[idx] is None:
                return False
            val = str(row[idx]).strip().lower()
            return val in ('true', '1', 'yes')

        records.append({
            'user_login': user_login,
            'day': day_str,
            'interactions': safe_int('interactions'),
            'code_gen': safe_int('code_gen'),
            'code_accept': safe_int('code_accept'),
            'used_agent': safe_bool('used_agent'),
            'used_chat': safe_bool('used_chat'),
            'loc_suggested': safe_int('loc_suggested'),
            'loc_added': safe_int('loc_added'),
        })

    return records


def aggregate_by_user(records, eas_usernames=None):
    """
    Aggregate daily records per user_login.
    If eas_usernames is provided, filter to only those users.
    """
    user_data = defaultdict(lambda: {
        'days_active': 0,
        'total_interactions': 0,
        'code_generations': 0,
        'code_acceptances': 0,
        'agent_days': 0,
        'chat_days': 0,
        'loc_suggested': 0,
        'loc_added': 0,
        'last_active_date': None,
        'first_date': None,
        'last_date': None,
        'days_set': set(),
    })

    for r in records:
        username = r['user_login']

        # Filter to EAS users if list provided
        if eas_usernames is not None and username not in eas_usernames:
            continue

        u = user_data[username]

        # Count unique active days
        if r['day'] and r['day'] not in u['days_set']:
            u['days_set'].add(r['day'])
            u['days_active'] += 1

        u['total_interactions'] += r['interactions']
        u['code_generations'] += r['code_gen']
        u['code_acceptances'] += r['code_accept']

        if r['used_agent']:
            u['agent_days'] += 1
        if r['used_chat']:
            u['chat_days'] += 1

        u['loc_suggested'] += r['loc_suggested']
        u['loc_added'] += r['loc_added']

        # Track date range
        if r['day']:
            if u['first_date'] is None or r['day'] < u['first_date']:
                u['first_date'] = r['day']
            if u['last_date'] is None or r['day'] > u['last_date']:
                u['last_date'] = r['day']
            if u['last_active_date'] is None or r['day'] > u['last_active_date']:
                u['last_active_date'] = r['day']

    # Clean up sets
    for username in user_data:
        del user_data[username]['days_set']

    return dict(user_data)


def gen_grafana_sql(user_aggregates):
    """Generate SQL UPDATEs for copilot_users based on Grafana aggregates."""
    lines = [
        "-- Grafana IDE Usage Sync (generated " + datetime.now().isoformat() + ")",
        "-- Updates copilot_users.ide_* columns by matching username",
        ""
    ]

    for username, data in sorted(user_aggregates.items()):
        period = ''
        if data['first_date'] and data['last_date']:
            period = f"{data['first_date']} to {data['last_date']}"

        lines.append(f"""UPDATE copilot_users SET
  ide_days_active = {data['days_active']},
  ide_total_interactions = {data['total_interactions']},
  ide_code_generations = {data['code_generations']},
  ide_code_acceptances = {data['code_acceptances']},
  ide_agent_days = {data['agent_days']},
  ide_chat_days = {data['chat_days']},
  ide_loc_suggested = {data['loc_suggested']},
  ide_loc_added = {data['loc_added']},
  ide_last_active_date = {escape_sql(data['last_active_date'])},
  ide_data_period = {escape_sql(period)},
  ide_data_updated_at = now()
WHERE username = {escape_sql(username)};
""")

    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='Sync Grafana IDE Usage → SQL')
    parser.add_argument('files', nargs='+', help='Grafana Excel file(s)')
    parser.add_argument('--emails', help='File with EAS employee emails (one per line) to filter')
    parser.add_argument('--out', default='scripts/sync_output', help='Output directory')
    args = parser.parse_args()

    os.makedirs(args.out, exist_ok=True)

    # Load EAS email list if provided
    eas_usernames = None
    if args.emails:
        with open(args.emails, 'r') as f:
            eas_usernames = set()
            for line in f:
                email = line.strip().lower()
                if '@' in email:
                    eas_usernames.add(email.split('@')[0])
                elif email:
                    eas_usernames.add(email)
        print(f"Loaded {len(eas_usernames)} EAS usernames from {args.emails}")

    # Read all Grafana files
    all_records = []
    for filepath in args.files:
        if not os.path.exists(filepath):
            print(f"WARNING: File not found: {filepath}")
            continue
        print(f"Reading: {filepath}")
        records = read_grafana_file(filepath)
        print(f"  {len(records)} daily records")
        all_records.extend(records)

    print(f"\nTotal daily records: {len(all_records)}")

    # Get unique usernames
    all_usernames = set(r['user_login'] for r in all_records)
    print(f"Unique user logins: {len(all_usernames)}")

    if eas_usernames:
        matched = all_usernames & eas_usernames
        print(f"EAS users matched: {len(matched)} of {len(eas_usernames)}")
        unmatched_eas = eas_usernames - all_usernames
        if unmatched_eas:
            print(f"  EAS users with no Grafana data: {sorted(unmatched_eas)[:10]}...")

    # Aggregate
    user_agg = aggregate_by_user(all_records, eas_usernames)
    print(f"Aggregated data for {len(user_agg)} users")

    # Generate SQL
    sql = gen_grafana_sql(user_agg)
    output_path = os.path.join(args.out, 'sync_grafana_ide.sql')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(sql)
    print(f"\n[OK] {output_path} ({len(user_agg)} UPDATE statements)")
    print("\nDONE! Execute via Supabase MCP execute_sql.")


if __name__ == '__main__':
    main()
